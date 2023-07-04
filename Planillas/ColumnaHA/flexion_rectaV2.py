# Datos columna
import numpy as np
import openpyxl
import utils

base=0.35
altura=0.5
#tension_carac=25
tension_carac=20.684

# Datos armadura
area= 4.91/10000
cant= [3,2,3]
area_total = area*sum(cant)
pos= [0.035,0.25,0.465]
Fy = 413.685

# Constantes
beta1= 0.85
ModYoung= 200000
Ecu=-0.003

# Conexion con Excel
wb = openpyxl.load_workbook('FlexionRecta.xlsx')
ws = wb.active
ws.cell(row=1, column=1, value = 'Pn')
ws.cell(row=1, column=2, value = 'Mn')
ws.cell(row=1, column=3, value = 'e')
ws.cell(row=1, column=4, value = 'c')
wb.save('FlexionRecta.xlsx')

# Calculo
c_array= np.arange(altura/beta1, -altura/2,-0.01)
row = 1

for c in c_array:
	phi= -Ecu/c
	Pcompresion= -0.85*tension_carac*beta1*(c*base-area_total)
	Pn= Pcompresion
	Mcompresion= Pcompresion*(beta1*c-altura)/2
	Mn= Mcompresion
	Totaltraccion = 0
	TotalMomentoTraccion = 0
	Es=[]
	for i in range(len(pos)):
		Esi= phi*(pos[i]-c)
		Es.append(Esi)
		if np.abs(Esi*ModYoung) < Fy: 
			fsi= Esi*ModYoung
		elif Esi*ModYoung < 0:
			fsi= -Fy
		else:
			fsi= Fy
	
		disty= pos[i]-altura*0.5
		Ptraccion= area*fsi*cant[i]
		Pn+= Ptraccion
		Totaltraccion+=Ptraccion
		Mtraccion=area*fsi*cant[i]*disty
		Mn+= Mtraccion
		TotalMomentoTraccion += Mtraccion
		if Pn > 0:
			Pn= Totaltraccion
			Mn= TotalMomentoTraccion

	e=-Mn/Pn	
	
	coef_reduccion = utils.coef_reduccion(np.max(Es))
	P0 = -0.8*(0.85*tension_carac*(base*altura-area_total)+area_total*Fy)
	if np.abs(Pn*coef_reduccion) > np.abs(P0*.65):
		Pn = P0

	if Mn>=0 and c>0:
		row+= 1
		if row == 2:
			ws.cell(row=row, column=1, value = -P0*0.65)
			ws.cell(row=row, column=2, value = 0)
			ws.cell(row=row, column=3, value = "-")
			ws.cell(row=row, column=4, value = "-")
		else:
			print(f"| Pn= {round(Pn,4)} | Mn= {round(Mn,4)} | e= {round(e,3)} | c={round(c,2)}")
			ws.cell(row=row, column=1, value = -Pn*coef_reduccion)
			ws.cell(row=row, column=2, value = Mn*coef_reduccion)
			ws.cell(row=row, column=3, value = e)
			ws.cell(row=row, column=4, value = c)

wb.save('FlexionRecta.xlsx')