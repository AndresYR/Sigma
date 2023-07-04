import numpy as np
import openpyxl

# Datos columna
base=0.35
altura=0.5
tension_carac=20.684
rec = .025

# Datos armadura
area= 4.91/10000
cant= [3,2,3]
pos= [0.035,0.25,0.465]
Fy = 551.58
diam_estribos = .006
diam_arm_esquinas = .025
diam_arm_bordes = .025
area_arm_esquinas = np.pi*diam_arm_esquinas^2/4
area_arm_bordes = np.pi*diam_arm_bordes^2/4
posicion = [
			# Esquina
			{x: rec + diam_estribos + diam_arm_esquinas/2,
			 y: altura-rec-diam_estribos-diam_arm_esquinas/2,
			 area: area_arm_esquinas},
			{x: base - (rec + diam_estribos + diam_arm_esquinas/2),
			 y: altura-rec-diam_estribos-diam_arm_esquinas/2,
			 area: area_arm_esquinas},
			{x: rec + diam_estribos + diam_arm_esquinas/2,
			 y: rec+diam_estribos+diam_arm_esquinas/2,
			 area: area_arm_esquinas},
			{x: base - (rec + diam_estribos + diam_arm_esquinas/2),
			 y: rec+diam_estribos+diam_arm_esquinas/2,
			 area: area_arm_esquinas},
			# Bordes
			{x: rec + diam_estribos + diam_arm_esquinas/2,
			 y: altura/2,
			 area: area_arm_bordes},
			{x: base - (rec + diam_estribos + diam_arm_esquinas/2),
			 y: altura/2,
			 area: area_arm_bordes},
			 ]

# Constantes
beta1= 0.85
ModYoung= 200000
Ecu=-0.003

# Conexion con Excel
wb = openpyxl.load_workbook('FlexionRecta.xlsx')
ws = wb.active
ws.cell(row=1, column=1, value = 'Pn')
ws.cell(row=1, column=2, value = 'Mn')
ws.cell(row=1, column=3, value = 'e')
ws.cell(row=1, column=4, value = 'c')
wb.save('FlexionRecta.xlsx')

# Calculo
c_array= np.arange(altura/beta1+.1, -altura/2,-0.01)
row = 1
for c in c_array:
	phi= -Ecu/c
	Pcompresion= -0.85*tension_carac*beta1*c*base
	Pn= Pcompresion
	Mcompresion=-0.85*tension_carac*beta1*c*base*(beta1*c-altura)/2
	Mn= Mcompresion

	for i in range(len(pos)):
		Esi= phi*(pos[i]-c)
		if np.abs(Esi*ModYoung) < Fy: 
			fsi= Esi*ModYoung
		elif Esi*ModYoung < 0:
			fsi= -Fy
		else:
			fsi= Fy
	
		disty= pos[i]-altura*0.5
		Ptraccion= area*fsi*cant[i]
		Pn+= Ptraccion
		Mtraccion=area*fsi*cant[i]*disty
		Mn+= Mtraccion
		if Pn > 0:
			Pn= Ptraccion
			Mn+= area*fsi*cant[i]*disty

	e=-Mn/Pn	
	
	if Mn>0:
		row+= 1
		print(f"| Pn= {round(Pn,4)} | Mn= {round(Mn,4)} | e= {round(e,3)} | c={round(c,2)}")
		ws.cell(row=row, column=1, value = -Pn)
		ws.cell(row=row, column=2, value = Mn)
		ws.cell(row=row, column=3, value = e)
		ws.cell(row=row, column=4, value = c)

wb.save('FlexionRecta.xlsx')